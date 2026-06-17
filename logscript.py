#!/usr/bin/env python3
"""
pkglog.py - Arch Linux package history tracker

Parses /var/log/pacman.log and writes a multi-sheet xlsx to ~/Scripts/pkglog.xlsx

Sheets:
  Explicitly Downloaded  - one row per explicit package + indented deps, sorted by last updated
  AUR                   - one row per AUR package + indented deps, sorted by last updated
  Official Dependencies  - base/orphan packages not pulled in by any explicit install
  History: Explicit      - flat chronological log for explicit packages
  History: AUR           - flat chronological log for AUR packages
  History: Official      - flat chronological log for official dependencies
  History                - full log of every pacman event

Usage:
  python3 pkglog.py           # generate / refresh the spreadsheet
  python3 pkglog.py --setup   # first-time setup: install hook + generate spreadsheet
"""

import re
import subprocess
import sys
import shutil
import argparse
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Install it with:  sudo pacman -S python-openpyxl")
    sys.exit(1)

SCRIPT_PATH = Path(__file__).resolve()
SCRIPT_DIR  = SCRIPT_PATH.parent
LOG_PATH    = Path("/var/log/pacman.log")
OUT_PATH    = SCRIPT_DIR / "pkglog.xlsx"
HOOK_DIR    = Path("/etc/pacman.d/hooks")
HOOK_PATH   = HOOK_DIR / "pkglog.hook"

HOOK_CONTENT = f"""\
[Trigger]
Operation = Install
Operation = Upgrade
Operation = Remove
Type = Package
Target = *

[Action]
Description = Updating pkglog package history...
When = PostTransaction
Exec = /usr/bin/python3 {SCRIPT_PATH}
"""

LOG_RE = re.compile(
    r"^\[(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2})[^\]]*\] \[ALPM\] "
    r"(installed|upgraded|removed|reinstalled) ([^\s]+) \((.+)\)$"
)

# header fills
EXPLICIT_FILL      = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
AUR_FILL           = PatternFill("solid", start_color="375623", end_color="375623")
OFFICIAL_FILL      = PatternFill("solid", start_color="4A235A", end_color="4A235A")
HIST_EXPLICIT_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
HIST_AUR_FILL      = PatternFill("solid", start_color="548235", end_color="548235")
HIST_OFFICIAL_FILL = PatternFill("solid", start_color="7030A0", end_color="7030A0")
HISTORY_FILL       = PatternFill("solid", start_color="7B3F00", end_color="7B3F00")

# row fills
PARENT_EXPLICIT_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
PARENT_AUR_FILL      = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
PARENT_OFFICIAL_FILL = PatternFill("solid", start_color="E2CFEE", end_color="E2CFEE")
DEP_FILL             = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
DEP_FONT    = Font(name="Arial", size=10, italic=True, color="666666")

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

ACTION_COLORS = {
    "installed":   "C6EFCE",
    "upgraded":    "DDEBF7",
    "reinstalled": "FFF2CC",
    "removed":     "FFCCCC",
}

SUMMARY_COLS   = ["Last Updated", "Package", "First Installed", "Total Updates", "Current Version"]
SUMMARY_WIDTHS = [16, 32, 16, 14, 20]
HISTORY_COLS   = ["Date", "Time", "Action", "Package", "Version / Change"]
HISTORY_WIDTHS = [14, 10, 14, 30, 40]


# ── setup ────────────────────────────────────────────────────────────────────

def check_dependencies():
    ok = True
    if not shutil.which("pacman"):
        print("Error: pacman not found. pkglog requires Arch Linux.")
        ok = False
    if not LOG_PATH.exists():
        print(f"Error: {LOG_PATH} not found.")
        ok = False
    return ok


def install_hook():
    if HOOK_PATH.exists():
        print(f"Hook already installed at {HOOK_PATH}")
        return True
    print(f"Installing pacman hook to {HOOK_PATH} (requires sudo)...")
    try:
        subprocess.run(["sudo", "mkdir", "-p", str(HOOK_DIR)], check=True)
        tmp = Path("/tmp/pkglog.hook")
        tmp.write_text(HOOK_CONTENT)
        subprocess.run(["sudo", "cp", str(tmp), str(HOOK_PATH)], check=True)
        subprocess.run(["sudo", "chmod", "644", str(HOOK_PATH)], check=True)
        tmp.unlink()
        print(f"Hook installed - {HOOK_PATH}")
        return True
    except subprocess.CalledProcessError:
        print("Failed to install hook. You can install it manually:")
        print(f"  sudo mkdir -p {HOOK_DIR}")
        print(f"  sudo tee {HOOK_PATH} << 'EOF'")
        print(HOOK_CONTENT + "EOF")
        return False


def setup():
    print("=== pkglog setup ===\n")
    if not check_dependencies():
        sys.exit(1)
    hook_ok = install_hook()
    print("\nGenerating initial spreadsheet...")
    run()
    print("\n=== Setup complete ===")
    print(f"Spreadsheet: {OUT_PATH}")
    if hook_ok:
        print(f"Hook:        {HOOK_PATH}")
        print("\nThe spreadsheet will auto-update after every pacman transaction.")
    else:
        print("\nHook installation failed - auto-update won't work until the hook is installed.")


# ── data gathering ────────────────────────────────────────────────────────────

def get_official_packages():
    try:
        r = subprocess.run(["pacman", "-Slq"], capture_output=True, text=True, timeout=15)
        return set(r.stdout.splitlines())
    except Exception:
        return set()


def get_explicit_packages():
    try:
        r = subprocess.run(["pacman", "-Qqe"], capture_output=True, text=True, timeout=15)
        return set(r.stdout.splitlines())
    except Exception:
        return set()


def batch_pkg_info(pkg_list):
    """
    Run a single pacman -Qi call for all packages and return a dict:
      { pkg_name: { "version": str, "deps": [str] } }
    """
    if not pkg_list:
        return {}
    try:
        r = subprocess.run(
            ["pacman", "-Qi"] + list(pkg_list),
            capture_output=True, text=True, timeout=60
        )
    except Exception:
        return {}

    info = {}
    current = None
    for line in r.stdout.splitlines():
        if line.startswith("Name"):
            current = line.split(":", 1)[1].strip()
            info[current] = {"version": "", "deps": []}
        elif current and line.startswith("Version"):
            info[current]["version"] = line.split(":", 1)[1].strip()
        elif current and line.startswith("Depends On"):
            val = line.split(":", 1)[1].strip()
            if val != "None":
                deps = re.split(r"\s+", val)
                info[current]["deps"] = [re.split(r"[><=]", d)[0] for d in deps if d]
    return info


# module-level cache populated once per run
_PKG_INFO_CACHE = {}


def get_deps_of(pkg):
    return _PKG_INFO_CACHE.get(pkg, {}).get("deps", [])


def get_current_version(pkg):
    return _PKG_INFO_CACHE.get(pkg, {}).get("version", "")


def parse_log():
    events = []
    with open(LOG_PATH, "r", errors="replace") as f:
        for line in f:
            m = LOG_RE.match(line.rstrip())
            if not m:
                continue
            date, time_, action, pkg, ver = m.groups()
            events.append({
                "date":     date,
                "time":     time_,
                "datetime": datetime.fromisoformat(f"{date}T{time_}"),
                "action":   action,
                "package":  pkg,
                "version":  ver,
            })
    return sorted(events, key=lambda e: e["datetime"])


def build_pkg_stats(events):
    """Build per-package stats dict from log events."""
    stats = defaultdict(lambda: {
        "first_installed": None,
        "last_updated":    None,
        "total_updates":   0,
        "last_action":     "installed",
        "events":          [],
    })
    for e in events:
        pkg = e["package"]
        s   = stats[pkg]
        s["events"].append(e)
        if e["action"] == "installed" and s["first_installed"] is None:
            s["first_installed"] = e["date"]
        if e["action"] in ("upgraded", "reinstalled"):
            s["total_updates"] += 1
            s["last_updated"] = e["date"]
        if s["first_installed"] and not s["last_updated"]:
            s["last_updated"] = s["first_installed"]
        s["last_action"] = e["action"]
    return stats


def classify_packages(pkg_stats, official_pkgs, explicit_pkgs):
    """
    Returns:
      explicit_set   - official repo packages explicitly installed by the user
      aur_set        - packages not in official repos (AUR/manual)
      system_set     - official repo packages NOT explicitly installed (auto deps)
    """
    explicit_set = set()
    aur_set      = set()
    system_set   = set()

    for pkg in pkg_stats:
        if pkg in official_pkgs:
            if pkg in explicit_pkgs:
                explicit_set.add(pkg)
            else:
                system_set.add(pkg)
        else:
            aur_set.add(pkg)

    return explicit_set, aur_set, system_set


# ── xlsx helpers ──────────────────────────────────────────────────────────────

def style_header(ws, headers, widths, fill):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_summary_row(ws, r, values, row_fill, font=None):
    for col, val in enumerate(values, 1):
        cell           = ws.cell(row=r, column=col, value=val)
        cell.font      = font or BODY_FONT
        cell.fill      = row_fill
        cell.border    = BORDER
        cell.alignment = Alignment(vertical="center")


# ── summary sheets ─────────────────────────────────────────────────────────────

def write_summary_sheet(ws, pkg_list, pkg_stats, header_fill, parent_fill, show_deps=True):
    """
    One row per package (sorted by last_updated desc), with dependency rows
    indented underneath each parent.
    """
    style_header(ws, SUMMARY_COLS, SUMMARY_WIDTHS, header_fill)

    # sort packages by last_updated descending
    def sort_key(pkg):
        lu = pkg_stats[pkg]["last_updated"]
        return lu if lu else "0000-00-00"

    sorted_pkgs = sorted(pkg_list, key=sort_key, reverse=True)

    r = 2
    for pkg in sorted_pkgs:
        s       = pkg_stats[pkg]
        version = get_current_version(pkg)
        lu      = s["last_updated"] or s["first_installed"] or ""
        fi      = s["first_installed"] or ""
        tu      = s["total_updates"]

        # parent row — color by last action
        action_color = ACTION_COLORS.get(s.get("last_action", "installed"), "C6EFCE")
        pkg_fill = PatternFill("solid", start_color=action_color, end_color=action_color)
        write_summary_row(ws, r, [lu, pkg, fi, tu, version], pkg_fill)
        r += 1

        if show_deps:
            deps = get_deps_of(pkg)
            for dep in sorted(deps):
                if dep not in pkg_stats:
                    continue
                ds           = pkg_stats[dep]
                dver         = get_current_version(dep)
                dlu          = ds["last_updated"] or ds["first_installed"] or ""
                dfi          = ds["first_installed"] or ""
                dtu          = ds["total_updates"]
                dep_action   = ds.get("last_action", "installed")
                dep_color    = ACTION_COLORS.get(dep_action, "C6EFCE")
                # blend dep color lighter by mixing with white
                r_hex = hex(min(255, int(dep_color[0:2], 16) + 40))[2:].zfill(2)
                g_hex = hex(min(255, int(dep_color[2:4], 16) + 40))[2:].zfill(2)
                b_hex = hex(min(255, int(dep_color[4:6], 16) + 40))[2:].zfill(2)
                light_color = r_hex + g_hex + b_hex
                dep_fill = PatternFill("solid", start_color=light_color, end_color=light_color)
                write_summary_row(ws, r, [dlu, f"  └ {dep}", dfi, dtu, dver],
                                  dep_fill, DEP_FONT)
                r += 1


# ── history sheets ─────────────────────────────────────────────────────────────

def write_history_sheet(ws, events, header_fill):
    style_header(ws, HISTORY_COLS, HISTORY_WIDTHS, header_fill)
    for r, e in enumerate(reversed(events), 2):
        c  = ACTION_COLORS.get(e["action"], "FFFFFF")
        rf = PatternFill("solid", start_color=c, end_color=c)
        for col, val in enumerate(
            [e["date"], e["time"], e["action"], e["package"], e["version"]], 1
        ):
            cell           = ws.cell(row=r, column=col, value=val)
            cell.font      = BODY_FONT
            cell.fill      = rf
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")


# ── main ───────────────────────────────────────────────────────────────────────

def run(include_history=True):
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    print("Reading official package list from pacman...")
    official_pkgs = get_official_packages()
    print(f"  {len(official_pkgs)} packages in official repos")

    print("Reading explicitly installed packages...")
    explicit_pkgs = get_explicit_packages()
    print(f"  {len(explicit_pkgs)} explicitly installed packages")

    print(f"Parsing {LOG_PATH}...")
    events    = parse_log()
    pkg_stats = build_pkg_stats(events)
    print(f"  {len(events)} events, {len(pkg_stats)} unique packages")

    print("Classifying packages...")
    explicit_set, aur_set, system_set = \
        classify_packages(pkg_stats, official_pkgs, explicit_pkgs)
    print(f"  {len(explicit_set)} explicit, {len(aur_set)} AUR, "
          f"{len(system_set)} system packages")

    print("Fetching package info (versions + dependencies)...")
    all_pkgs = set(pkg_stats.keys())
    global _PKG_INFO_CACHE
    _PKG_INFO_CACHE = batch_pkg_info(all_pkgs)
    print(f"  {len(_PKG_INFO_CACHE)} packages queried")

    wb = Workbook()

    # ── summary sheets ──
    ws_explicit = wb.active
    ws_explicit.title = "Official Repository"
    write_summary_sheet(ws_explicit, explicit_set, pkg_stats,
                        EXPLICIT_FILL, PARENT_EXPLICIT_FILL, show_deps=True)
    c = Comment(
        "Removed packages will not appear here — pacman -Qqe only tracks currently "
        "installed packages. Check History Official Repo for removed package events.",
        "pkglog"
    )
    ws_explicit["A1"].comment = c

    ws_aur = wb.create_sheet("AUR")
    write_summary_sheet(ws_aur, aur_set, pkg_stats,
                        AUR_FILL, PARENT_AUR_FILL, show_deps=True)

    ws_system = wb.create_sheet("System Packages")
    write_summary_sheet(ws_system, system_set, pkg_stats,
                        OFFICIAL_FILL, PARENT_OFFICIAL_FILL, show_deps=False)

    # ── history sheets ──
    if include_history:
        explicit_events = [e for e in events if e["package"] in explicit_set]
        aur_events      = [e for e in events if e["package"] in aur_set]
        official_events = [e for e in events if e["package"] in system_set]

        ws_hist_explicit = wb.create_sheet("History Official Repo")
        write_history_sheet(ws_hist_explicit, explicit_events, HIST_EXPLICIT_FILL)

        ws_hist_aur = wb.create_sheet("History AUR")
        write_history_sheet(ws_hist_aur, aur_events, HIST_AUR_FILL)

        ws_hist_official = wb.create_sheet("History System")
        write_history_sheet(ws_hist_official, official_events, HIST_OFFICIAL_FILL)

    ws_history = wb.create_sheet("History")
    write_history_sheet(ws_history, events, HISTORY_FILL)

    wb.save(OUT_PATH)
    print(f"Saved - {OUT_PATH}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="pkglog - Arch Linux package history tracker"
    )
    parser.add_argument(
        "--setup",
        action="store_true",
        help="First-time setup: install pacman hook and generate spreadsheet"
    )
    args = parser.parse_args()

    if args.setup:
        setup()
    else:
        if not check_dependencies():
            sys.exit(1)
        print("""
Row colors in the spreadsheet:
  Green  - installed   (package was freshly installed)
  Blue   - upgraded    (package was upgraded to a newer version)
  Yellow - reinstalled (package was reinstalled at the same version)
  Red    - removed     (package was removed from the system)

Dependency rows appear indented under their parent package
and use a lighter shade of the same color.
""")
        ans = input("Generate per-category history sheets? This will take more time. [y/N]: ").strip().lower()
        run(include_history=(ans == "y"))
